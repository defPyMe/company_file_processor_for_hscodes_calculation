import openpyxl
from openpyxl import Workbook


def writing_to_file_f(file, value):
    scfile = openpyxl.load_workbook(file, read_only=False)
    ws = scfile.active
    for i in value:#single list
        ws.cell(row=int(i[1]), column = 1 ).value = i[7]
        ws.cell(row=int(i[1]), column = 2 ).value = i[8]
        ws.cell(row=int(i[1]), column = 3 ).value = i[2]
        ws.cell(row=int(i[1]), column = 4 ).value = i[5]
        ws.cell(row=int(i[1]), column = 5 ).value = i[6]
        ws.cell(row=int(i[1]), column = 6 ).value = i[4]
        ws.cell(row=int(i[1]), column = 7 ).value = i[3]
        ws.cell(row=int(i[1]), column = 8 ).value = i[0]

    scfile.save(r'C:\Users\cavazzinil\Dropbox\hs_codes_calculator\all_excel_files\balances.xlsx')